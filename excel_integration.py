from openpyxl import load_workbook
from openpyxl.utils import get_column_letter, column_index_from_string

class ExcelIntegration():
    def __init__(self):
        self.requirements = 'openpyxl'

    def set_input_params(self, file_path, sheet_name, table_head = 1):
        self.file_path = file_path
        self.sheet_name = sheet_name
        self.workbook = load_workbook(filename=file_path)
        self.sheet = self.workbook[sheet_name]
        self.table_head = table_head

    def __get_cells(self, row_index, colnames):
        row = self.__row_index_convert(row_index)
        columns = self.__colname_2_index(colnames)
        if len(columns) == 1:
            return self.sheet.cell(row=row, column=columns[0])
        else:
            return [self.sheet.cell(row=row, column=column) for column in range(columns[0], columns[1]+1)]

    def __single_cell_2_record(self,row, colnames, key, mode = 'None'):
        cells = self.__get_cells(row, colnames)
        if isinstance(cells, list):
            result = {key: cells[0].value}
        else:
            result = {key: cells.value}

        if mode == 'Boolen':
            result[key] = bool(result[key])
        elif mode == 'int':
            result[key] = int(result[key])
        return result

    def __multi_cells_2_record(self,row, colnames, keys='PublishYear', *category,):
        cells = self.__get_cells(row, colnames)
        if keys == 'PublishYear':
            year = 1999
            for cell in cells:
                if cell.value == 1:
                    return {keys: year}  
                year += 1
        elif len(keys)>1 and len(keys)==self.__column_distance(colnames):
            result = {}
            for key, cell in zip(keys, cells):
                singleresult = self.__single_cell_2_record(row, cell.column_letter, key, mode = 'Boolen')
                result.update(singleresult)
            return {category[0]:result}
        else:
            print('method not implemented yet.')
            return None

    def cells_2_record(self, row, colnames, keys='PublishYear', *category, Type = 'None',mode='None'):
        if Type == 'single':
            return self.__single_cell_2_record(row, colnames, keys, mode)
        elif Type == 'multi':
            return self.__multi_cells_2_record(row, colnames, keys, *category)
        else:
            raise ValueError('cell2record mode should be single or multi, please check it.')
            return None

    def __row_index_convert(self, indexs):
        return indexs+self.table_head

    def read_one_articlle(self,row):
        article = {}
        from excel_template import data_extraction_operations
        self.data_extraction_operations = data_extraction_operations

        # Process each specification and update the article dictionary
        for operation in self.data_extraction_operations:
            colnames = operation[1]
            Type = operation[0]
            keys = operation[2]
            category, mode= None, None 
            if len(operation) > 3:
                category = operation[3]
                if len(operation) > 4:
                    mode = operation[4]

            result = self.cells_2_record(row, colnames, keys, category, Type = Type, mode = mode)
            article.update(result)
        
        # add title from citation if not found
        if 'Title' not in article and 'Full_Citation' in article:

            article['Title'] = self.__extract_title_from_citation(article['Full_Citation'])

        article.update({'import_source': 'excel'})
        return article
    
    def read_articles(self, start_id, end_id):
        articles = []
        for row in range(start_id, end_id+1):
            article = self.read_one_articlle(row)
            articles.append(article)
        return articles
    
    @staticmethod   
    def __column_distance(colnames):
        return column_index_from_string(colnames[1])-column_index_from_string(colnames[0])+1

    @staticmethod
    def __colname_2_index(colnames):
        if isinstance(colnames, str):
            colnames = [colnames]
        return [column_index_from_string(colname) for colname in colnames]
    
    @staticmethod
    def __extract_title_from_citation(citation):
        # extract title from citation
        patterns = [
            r'\(\d{4}\)\s(.*?)\.',         # Format: (year) Title. Other fields
            r'\(\d{4}\)\.\s“(.*?)\.”',     # Format: (year). "Title." Other fields
            r'\(\d{4}\)\.\s(.*?)\.\s'      # Format: (year). Title. Other fields
        ]
        for pattern in patterns:
            match = re.search(pattern, citation)
            if match:
                return match.group(1)
        return None